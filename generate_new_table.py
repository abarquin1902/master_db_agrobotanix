# python3 generate_new_table.py

import openpyxl
import pandas as pd

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Rutas ────────────────────────────────────────────────────────────────────
EXCEL_INPUT  = "master_db_agrobotanix.xlsx"
EXCEL_OUTPUT = "db_agrobotanix.xlsx"

# ─── Orden de dificultad ──────────────────────────────────────────────────────
ORDEN_DIF = {"alta": 0, "media-alta": 1, "media": 2, "baja-media": 3, "baja": 4}

TIPOS = ["virus", "bacterias", "hongos"]

LABELS_TIPO = {
    "virus":     "🦠 VIRUS",
    "bacterias": "🔬 BACTERIAS",
    "hongos":    "🍄 HONGOS",
}

# ─── Colores ──────────────────────────────────────────────────────────────────
COLOR_HEADER_ESTADO  = "1A3C2E"   # verde oscuro — encabezado principal
COLOR_HEADER_CULTIVO = "2D6A4F"   # verde medio — fila de cultivo
COLOR_VIRUS          = "3D1F6E"   # morado oscuro
COLOR_BACTERIAS      = "1A3A5C"   # azul oscuro
COLOR_HONGOS         = "4A1942"   # morado rojizo
COLOR_SUBHEADER      = "0A1628"   # fondo oscuro labels #1 #2 #3
COLOR_ROW_ALT        = "111F30"   # fila alternada
COLOR_FILA_VACIA     = "0D1B2A"   # celda sin datos

FONT_WHITE  = Font(name="Arial", color="FFFFFF", bold=True, size=9)
FONT_NORMAL = Font(name="Arial", color="D4E6F1", size=8)
FONT_ITALIC = Font(name="Arial", color="7FB3D3", italic=True, size=8)
FONT_DIM    = Font(name="Arial", color="2D5A7B", size=8)

THIN = Side(style="thin", color="1E3A5F")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def norm_dif(v):
    if pd.isna(v):
        return "media"
    v = str(v).strip().lower()
    if v in ("alta", "alto"):                              return "alta"
    if v in ("media-alta", "media–alta", "medio-alto"):   return "media-alta"
    if v in ("media", "medio"):                            return "media"
    if v in ("baja-media", "baja–media", "bajo-medio"):   return "baja-media"
    if v in ("baja", "bajo"):                              return "baja"
    return "media"

def top3_por_tipo(df_cultivo, tipo):
    """Devuelve lista de hasta 3 dicts con nombre_comun y nombre_cientifico."""
    subset = df_cultivo[df_cultivo["tipo_patogeno"] == tipo].copy()
    subset["orden"] = subset["dificultad_norm"].map(lambda x: ORDEN_DIF.get(x, 99))
    subset = subset.sort_values("orden")
    return subset[["nombre_comun", "nombre_cientifico"]].head(3).to_dict("records")

def apply_fill(cell, hex_color):
    cell.fill = PatternFill("solid", start_color=hex_color)

def apply_border(cell):
    cell.border = BORDER

def write_cell(ws, row, col, value, font, fill_color, alignment=ALIGN_LEFT, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    apply_fill(c, fill_color)
    c.alignment = alignment
    if border:
        apply_border(c)
    return c

def main():
    # ── Cargar datos ──────────────────────────────────────────────────────────
    xl  = pd.read_excel(EXCEL_INPUT, sheet_name=None)
    enf = xl["tbl_enfermedades"].copy()
    cal = xl["tbl_cultivo_calendario"][["estado", "cultivo"]].drop_duplicates()

    enf["dificultad_norm"] = enf["dificultad_erradicacion"].apply(norm_dif)

    # Top 3 por cultivo × tipo (precalculado una vez)
    top3_cache = {}
    for cultivo in enf["cultivo"].unique():
        df_c = enf[enf["cultivo"] == cultivo]
        top3_cache[cultivo] = {t: top3_por_tipo(df_c, t) for t in TIPOS}

    # Estados ordenados alfabéticamente
    estados = sorted(cal["estado"].unique())

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enfermedades por Estado"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # ─────────────────────────────────────────────────────────────────────────
    # ESTRUCTURA DE COLUMNAS
    # Col 1: Estado
    # Col 2: Cultivo
    # Por cada tipo (3): 3 sub-columnas → Nombre común + Nombre científico
    # = 2 + 3*3*2 = 2 + 18 = 20 columnas totales
    #
    # Mapa de columnas:
    #   A  = Estado
    #   B  = Cultivo
    #   C,D  = Virus #1 (nombre común, nombre científico)
    #   E,F  = Virus #2
    #   G,H  = Virus #3
    #   I,J  = Bacterias #1
    #   K,L  = Bacterias #2
    #   M,N  = Bacterias #3
    #   O,P  = Hongos #1
    #   Q,R  = Hongos #2
    #   S,T  = Hongos #3
    # ─────────────────────────────────────────────────────────────────────────

    COL_ESTADO  = 1
    COL_CULTIVO = 2

    # Inicio de columna por tipo
    COL_START = {"virus": 3, "bacterias": 9, "hongos": 15}

    # ── Fila 1: encabezados de grupo por tipo ─────────────────────────────────
    # Estado y Cultivo ocupan 2 filas (merge vertical)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    c = ws.cell(row=1, column=1, value="ESTADO")
    c.font = FONT_WHITE
    apply_fill(c, COLOR_HEADER_ESTADO)
    c.alignment = ALIGN_CENTER
    apply_border(c)

    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    c = ws.cell(row=1, column=2, value="CULTIVO")
    c.font = FONT_WHITE
    apply_fill(c, COLOR_HEADER_ESTADO)
    c.alignment = ALIGN_CENTER
    apply_border(c)

    # Encabezados de tipo (merge 6 columnas cada uno: 3 enfermedades × 2 cols)
    COLOR_TIPO = {
        "virus":     COLOR_VIRUS,
        "bacterias": COLOR_BACTERIAS,
        "hongos":    COLOR_HONGOS,
    }
    for tipo in TIPOS:
        cs = COL_START[tipo]
        ws.merge_cells(start_row=1, start_column=cs, end_row=1, end_column=cs + 5)
        c = ws.cell(row=1, column=cs, value=LABELS_TIPO[tipo])
        c.font = FONT_WHITE
        apply_fill(c, COLOR_TIPO[tipo])
        c.alignment = ALIGN_CENTER
        apply_border(c)
        # Rellenar celdas mergeadas (openpyxl requiere formato en cada celda)
        for col in range(cs + 1, cs + 6):
            cc = ws.cell(row=1, column=col)
            apply_fill(cc, COLOR_TIPO[tipo])
            apply_border(cc)

    # ── Fila 2: sub-encabezados #1 #2 #3 × (Nombre / Científico) ─────────────
    for tipo in TIPOS:
        cs = COL_START[tipo]
        for rank in range(3):
            base = cs + rank * 2
            # Merge 2 columnas para el label de rank
            ws.merge_cells(start_row=2, start_column=base, end_row=2, end_column=base + 1)
            c = ws.cell(row=2, column=base, value=f"#{rank+1}")
            c.font = Font(name="Arial", color="74C69D", bold=True, size=8)
            apply_fill(c, COLOR_SUBHEADER)
            c.alignment = ALIGN_CENTER
            apply_border(c)
            cc = ws.cell(row=2, column=base + 1)
            apply_fill(cc, COLOR_SUBHEADER)
            apply_border(cc)

    # ── Anchos de columna ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 16   # Estado
    ws.column_dimensions["B"].width = 22   # Cultivo
    # Para cada enfermedad: columna nombre común (más ancha) + científico
    col_letters = [get_column_letter(i) for i in range(3, 21)]
    for i, letter in enumerate(col_letters):
        ws.column_dimensions[letter].width = 28 if i % 2 == 0 else 26

    # ── Filas de datos ────────────────────────────────────────────────────────
    current_row = 3

    for estado in estados:
        cultivos_estado = sorted(cal[cal["estado"] == estado]["cultivo"].tolist())
        n_cultivos = len(cultivos_estado)

        # Merge del estado sobre todos sus cultivos
        if n_cultivos > 1:
            ws.merge_cells(
                start_row=current_row, start_column=COL_ESTADO,
                end_row=current_row + n_cultivos - 1, end_column=COL_ESTADO
            )

        # Escribir estado en la celda inicial (merge lo expande)
        c_estado = ws.cell(row=current_row, column=COL_ESTADO,
                           value=estado.title())
        c_estado.font = FONT_WHITE
        apply_fill(c_estado, COLOR_HEADER_CULTIVO)
        c_estado.alignment = ALIGN_CENTER
        apply_border(c_estado)

        # Rellenar celdas mergeadas del estado con el mismo color
        for r in range(current_row + 1, current_row + n_cultivos):
            cc = ws.cell(row=r, column=COL_ESTADO)
            apply_fill(cc, COLOR_HEADER_CULTIVO)
            apply_border(cc)

        for idx, cultivo in enumerate(cultivos_estado):
            row = current_row + idx
            fill_color = COLOR_ROW_ALT if idx % 2 == 0 else "0F1E2E"

            # Columna cultivo
            c = ws.cell(row=row, column=COL_CULTIVO, value=cultivo.title())
            c.font = Font(name="Arial", color="95D5B2", bold=True, size=8)
            apply_fill(c, fill_color)
            c.alignment = ALIGN_LEFT
            apply_border(c)

            # Top 3 por tipo
            for tipo in TIPOS:
                cs = COL_START[tipo]
                enfermedades = top3_cache.get(cultivo, {}).get(tipo, [])

                for rank in range(3):
                    base = cs + rank * 2
                    if rank < len(enfermedades):
                        enf = enfermedades[rank]
                        nombre_com = enf["nombre_comun"] or ""
                        nombre_cie = enf["nombre_cientifico"] or ""
                    else:
                        nombre_com = ""
                        nombre_cie = ""

                    # Nombre común
                    c1 = ws.cell(row=row, column=base, value=nombre_com if nombre_com else "—")
                    c1.font = FONT_NORMAL if nombre_com else FONT_DIM
                    apply_fill(c1, fill_color if nombre_com else COLOR_FILA_VACIA)
                    c1.alignment = ALIGN_LEFT
                    apply_border(c1)

                    # Nombre científico
                    c2 = ws.cell(row=row, column=base + 1, value=nombre_cie if nombre_cie else "")
                    c2.font = FONT_ITALIC if nombre_cie else FONT_DIM
                    apply_fill(c2, fill_color if nombre_cie else COLOR_FILA_VACIA)
                    c2.alignment = ALIGN_LEFT
                    apply_border(c2)

        # Altura de filas
        for r in range(current_row, current_row + n_cultivos):
            ws.row_dimensions[r].height = 28

        current_row += n_cultivos

    # ── Altura encabezados ────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16

    # ── Tab color ─────────────────────────────────────────────────────────────
    ws.sheet_properties.tabColor = "2D6A4F"

    wb.save(EXCEL_OUTPUT)
    print(f"✅ Archivo generado: {EXCEL_OUTPUT}")
    print(f" {len(estados)} estados · {len(cal)} combinaciones estado-cultivo")

if __name__ == "__main__":
    main()